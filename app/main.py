import json
from io import BytesIO
import os
from pathlib import Path
import re
import sqlite3
import sys
from datetime import datetime
import time
from typing import Any, TYPE_CHECKING

from fastapi import FastAPI, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font
import numpy as np
from PIL import Image, ImageFilter, ImageOps, UnidentifiedImageError
from pydantic import BaseModel

if TYPE_CHECKING:
    from paddleocr import PaddleOCR


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "plant_label_helper.db"
APP_HOME = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else BASE_DIR.parent
)
BACKUP_DIR = APP_HOME / "backup"
ENV_PATH = APP_HOME / ".env"
TURSO_CONFIG_PATH = APP_HOME / "turso_config.json"
TURSO_REPLICA_DIR = APP_HOME / "data" / "turso"
TURSO_REPLICA_PATH = TURSO_REPLICA_DIR / "purchase_entries_replica.db"
TURSO_URL_ENV_KEY = "TURSO_DATABASE_URL"
TURSO_TOKEN_ENV_KEY = "TURSO_AUTH_TOKEN"
TURSO_READ_SYNC_INTERVAL_SECONDS = 2.0
LAST_TURSO_SYNC_AT = 0.0

app = FastAPI(title="Plant Label Helper")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


class PlantRow(BaseModel):
    category: str = "plant"
    name: str = ""
    vendor: str = ""
    spec: str = ""
    cost: int | float | str = ""
    wholesale: int | float | str = ""
    retail: int | float | str = ""
    quantity: int | float | str = ""
    purchase_count: int | float | str = ""


class PurchaseLedgerRowUpdate(BaseModel):
    category: str = "plant"
    name: str = ""
    vendor: str = ""
    spec: str = ""
    quantity: int | float | str = ""
    purchase_count: int | float | str = ""
    cost: int | float | str = ""
    wholesale: int | float | str = ""
    retail: int | float | str = ""


class PurchaseLedgerExportSelection(BaseModel):
    ids: list[int] = []


class TursoConfigPayload(BaseModel):
    url: str = ""
    auth_token: str = ""


NUMERIC_WITH_COMMAS = re.compile(r"^\d{1,3}(,\d{3})+$")
NUMERIC_PLAIN_MONEY = re.compile(r"^\d{4,7}$")
PURE_NUMBER = re.compile(r"^\d+$")
SPEC_TOKEN = re.compile(r"^\d+(?:/\d+)?$")
VALID_CATEGORIES = {"plant", "material"}
OCR_ENGINE: Any | None = None


def get_ocr_engine() -> "PaddleOCR":
    global OCR_ENGINE
    if OCR_ENGINE is None:
        from paddleocr import PaddleOCR

        OCR_ENGINE = PaddleOCR(
            use_angle_cls=True,
            lang="korean",
            use_gpu=False,
            show_log=False,
        )
    return OCR_ENGINE


def load_saved_turso_config() -> dict[str, str] | None:
    if not TURSO_CONFIG_PATH.exists():
        return None

    try:
        payload = json.loads(TURSO_CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None

    url = str(payload.get("url", "")).strip()
    auth_token = str(payload.get("auth_token", "")).strip()
    if not url or not auth_token:
        return None

    return {"url": url, "auth_token": auth_token}


def parse_dotenv_value(raw_value: str) -> str:
    value = raw_value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        return value[1:-1]
    return value


def load_dotenv_values() -> dict[str, str]:
    if not ENV_PATH.exists():
        return {}

    try:
        lines = ENV_PATH.read_text(encoding="utf-8").splitlines()
    except OSError:
        return {}

    values: dict[str, str] = {}
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        if stripped.startswith("export "):
            stripped = stripped[len("export ") :].strip()
        if "=" not in stripped:
            continue

        key, raw_value = stripped.split("=", 1)
        key = key.strip()
        if not key:
            continue
        values[key] = parse_dotenv_value(raw_value)

    return values


def load_environment_turso_config() -> dict[str, str] | None:
    dotenv_values = load_dotenv_values()
    process_url = str(os.getenv(TURSO_URL_ENV_KEY) or "").strip()
    process_token = str(os.getenv(TURSO_TOKEN_ENV_KEY) or "").strip()
    dotenv_url = str(dotenv_values.get(TURSO_URL_ENV_KEY, "")).strip()
    dotenv_token = str(dotenv_values.get(TURSO_TOKEN_ENV_KEY, "")).strip()

    url = process_url or dotenv_url
    auth_token = process_token or dotenv_token
    if not url or not auth_token:
        return None

    if process_url or process_token:
        source = "process-env" if process_url and process_token else "mixed-env"
    else:
        source = "dotenv"

    return {"url": url, "auth_token": auth_token, "source": source}


def save_turso_config(url: str, auth_token: str) -> None:
    TURSO_CONFIG_PATH.write_text(
        json.dumps(
            {"url": url.strip(), "auth_token": auth_token.strip()},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def get_active_turso_config() -> dict[str, str] | None:
    environment_config = load_environment_turso_config()
    if environment_config:
        return environment_config

    saved_config = load_saved_turso_config()
    if saved_config:
        return {**saved_config, "source": "saved-config"}

    return None


def require_libsql() -> Any:
    try:
        import libsql
    except ImportError as error:
        raise RuntimeError(
            "Turso 연결에 필요한 libsql 패키지가 설치되지 않았습니다. requirements 설치 후 다시 실행해 주세요."
        ) from error

    return libsql


def close_connection(connection: Any | None) -> None:
    if connection is None:
        return

    close = getattr(connection, "close", None)
    if callable(close):
        close()


def maybe_sync_connection(connection: Any) -> None:
    sync = getattr(connection, "sync", None)
    if callable(sync):
        sync()


def mark_turso_sync_completed() -> None:
    global LAST_TURSO_SYNC_AT
    LAST_TURSO_SYNC_AT = time.monotonic()


def should_sync_turso_replica(force: bool = False) -> bool:
    if force or not TURSO_REPLICA_PATH.exists():
        return True

    return (time.monotonic() - LAST_TURSO_SYNC_AT) >= TURSO_READ_SYNC_INTERVAL_SECONDS


def cursor_to_dicts(cursor: Any) -> list[dict[str, Any]]:
    columns = [str(column[0]) for column in (cursor.description or [])]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def cursor_to_dict(cursor: Any) -> dict[str, Any] | None:
    columns = [str(column[0]) for column in (cursor.description or [])]
    row = cursor.fetchone()
    if row is None:
        return None
    return dict(zip(columns, row))


def fetch_scalar(connection: Any, query: str, params: tuple[Any, ...] = ()) -> Any:
    cursor = connection.execute(query, params)
    row = cursor.fetchone()
    if row is None:
        return None
    if isinstance(row, (tuple, list)):
        return row[0]
    return row


def ensure_purchase_entries_schema(connection: Any) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL DEFAULT 'plant',
            name TEXT NOT NULL,
            vendor TEXT,
            spec TEXT,
            quantity REAL,
            purchase_count REAL,
            cost REAL,
            wholesale REAL,
            retail REAL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    existing_columns = {
        str(row[1]) for row in connection.execute("PRAGMA table_info(purchase_entries)").fetchall()
    }
    if "vendor" not in existing_columns:
        connection.execute("ALTER TABLE purchase_entries ADD COLUMN vendor TEXT")
    if "purchase_count" not in existing_columns:
        connection.execute("ALTER TABLE purchase_entries ADD COLUMN purchase_count REAL")
    if "category" not in existing_columns:
        connection.execute(
            "ALTER TABLE purchase_entries ADD COLUMN category TEXT NOT NULL DEFAULT 'plant'"
        )
    connection.execute(
        """
        UPDATE purchase_entries
        SET category = 'plant'
        WHERE category IS NULL OR TRIM(category) = ''
        """
    )
    connection.commit()
    maybe_sync_connection(connection)


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as connection:
        ensure_purchase_entries_schema(connection)


init_db()


def get_database_settings_payload() -> dict[str, Any]:
    config = get_active_turso_config()
    if config:
        source = config.get("source", "saved-config")
        source_label_map = {
            "dotenv": ".env 파일",
            "process-env": "시스템 환경변수",
            "mixed-env": "시스템 환경변수 + .env",
            "saved-config": "앱 저장 설정",
        }
        return {
            "mode": "turso",
            "configured": True,
            "url": config["url"],
            "token_saved": True,
            "backup_dir": str(BACKUP_DIR),
            "storage_label": f"Turso ({config['url']})",
            "config_source": source,
            "config_source_label": source_label_map.get(source, "앱 저장 설정"),
            "env_path": str(ENV_PATH),
        }

    return {
        "mode": "local",
        "configured": False,
        "url": "",
        "token_saved": False,
        "backup_dir": str(BACKUP_DIR),
        "storage_label": str(DB_PATH),
        "config_source": "local",
        "config_source_label": "로컬 SQLite",
        "env_path": str(ENV_PATH),
    }


def sync_turso_replica(force: bool = False) -> None:
    if not should_sync_turso_replica(force):
        return

    config = get_active_turso_config()
    if not config:
        return

    TURSO_REPLICA_DIR.mkdir(parents=True, exist_ok=True)
    libsql = require_libsql()
    connection = None
    try:
        connection = libsql.connect(
            str(TURSO_REPLICA_PATH),
            sync_url=config["url"],
            auth_token=config["auth_token"],
        )
        maybe_sync_connection(connection)
        ensure_purchase_entries_schema(connection)
        mark_turso_sync_completed()
    except Exception as error:
        close_connection(connection)
        raise RuntimeError(f"Turso replica sync failed: {error}") from error
    finally:
        close_connection(connection)


def open_purchase_connection(force_sync: bool = False, for_write: bool = False) -> tuple[Any, str]:
    config = get_active_turso_config()
    if not config:
        connection = sqlite3.connect(DB_PATH)
        ensure_purchase_entries_schema(connection)
        return connection, "local"

    if not for_write:
        sync_turso_replica(force=force_sync)
        connection = sqlite3.connect(TURSO_REPLICA_PATH)
        ensure_purchase_entries_schema(connection)
        return connection, "turso"

    TURSO_REPLICA_DIR.mkdir(parents=True, exist_ok=True)
    libsql = require_libsql()
    connection = None
    try:
        connection = libsql.connect(
            str(TURSO_REPLICA_PATH),
            sync_url=config["url"],
            auth_token=config["auth_token"],
        )
        maybe_sync_connection(connection)
        mark_turso_sync_completed()
        ensure_purchase_entries_schema(connection)
        return connection, "turso"
    except Exception as error:
        close_connection(connection)
        raise RuntimeError(f"Turso 데이터베이스에 연결하지 못했습니다: {error}") from error


def read_local_purchase_rows() -> list[tuple[Any, ...]]:
    if not DB_PATH.exists():
        return []

    with sqlite3.connect(DB_PATH) as connection:
        return connection.execute(
            """
            SELECT category,
                   name,
                   vendor,
                   spec,
                   quantity,
                   purchase_count,
                   cost,
                   wholesale,
                   retail
            FROM purchase_entries
            ORDER BY id ASC
            """
        ).fetchall()


def maybe_migrate_local_rows_to_remote(connection: Any) -> int:
    remote_count = int(fetch_scalar(connection, "SELECT COUNT(*) FROM purchase_entries") or 0)
    if remote_count > 0:
        return 0

    local_rows = read_local_purchase_rows()
    if not local_rows:
        return 0

    connection.executemany(
        """
        INSERT INTO purchase_entries (
            category,
            name,
            vendor,
            spec,
            quantity,
            purchase_count,
            cost,
            wholesale,
            retail
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        local_rows,
    )
    connection.commit()
    maybe_sync_connection(connection)
    return len(local_rows)


def build_backup_path() -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    date_label = datetime.now().strftime("%Y-%m-%d")
    candidate = BACKUP_DIR / f"purchase-ledger-{date_label}.db"
    if not candidate.exists():
        return candidate

    index = 2
    while True:
        numbered = BACKUP_DIR / f"purchase-ledger-{date_label}-{index}.db"
        if not numbered.exists():
            return numbered
        index += 1


def cleanup_replica_sidecars(path: Path) -> None:
    for candidate in (path, path.with_suffix(path.suffix + "-wal"), path.with_suffix(path.suffix + "-shm")):
        if candidate.exists():
            try:
                candidate.unlink()
            except OSError:
                pass


def create_database_backup() -> tuple[Path, str]:
    backup_path = build_backup_path()
    config = get_active_turso_config()
    if not config:
        if not DB_PATH.exists():
            init_db()
        with sqlite3.connect(DB_PATH) as source_connection, sqlite3.connect(backup_path) as target_connection:
            source_connection.backup(target_connection)
        return backup_path, "local"

    replica_path = BACKUP_DIR / ".purchase-ledger-backup-replica.db"
    cleanup_replica_sidecars(replica_path)
    connection = None
    try:
        libsql = require_libsql()
        connection = libsql.connect(
            str(replica_path),
            sync_url=config["url"],
            auth_token=config["auth_token"],
        )
        maybe_sync_connection(connection)
    except Exception as error:
        raise RuntimeError(f"Turso 백업 파일을 준비하지 못했습니다: {error}") from error
    finally:
        close_connection(connection)

    with sqlite3.connect(replica_path) as source_connection, sqlite3.connect(backup_path) as target_connection:
        source_connection.backup(target_connection)

    cleanup_replica_sidecars(replica_path)
    return backup_path, "turso"


def normalize_text(value: str) -> str:
    return (
        str(value)
        .replace("O", "0")
        .replace("o", "0")
        .replace("l", "1")
        .replace("|", "1")
        .strip()
    )


def parse_money_token(token: str) -> int | None:
    cleaned = normalize_text(token).replace(" ", "")
    if NUMERIC_WITH_COMMAS.match(cleaned):
        return int(cleaned.replace(",", ""))
    if NUMERIC_PLAIN_MONEY.match(cleaned):
        return int(cleaned)
    return None


def parse_quantity_token(token: str) -> int | None:
    cleaned = normalize_text(token)
    if not PURE_NUMBER.match(cleaned):
        return None
    if len(cleaned) > 3:
        return None
    return int(cleaned)


def is_name_token(token: str) -> bool:
    cleaned = normalize_text(token)
    if not cleaned:
        return False
    if SPEC_TOKEN.match(cleaned):
        return False
    if NUMERIC_WITH_COMMAS.match(cleaned):
        return False
    if PURE_NUMBER.match(cleaned):
        return False
    return bool(re.search(r"[가-힣A-Za-z]", cleaned))


def parse_spec_divisor(spec: str) -> int | None:
    cleaned = normalize_text(spec)
    match = re.match(r"^(\d+)(?:/\d+)?$", cleaned)
    if not match:
        return None
    value = int(match.group(1))
    return value if value > 0 else None


def parse_number(value: int | float | str) -> int | float | str:
    if isinstance(value, (int, float)):
        return value
    cleaned = normalize_text(value).replace(",", "")
    if not cleaned:
        return ""
    return float(cleaned) if cleaned.replace(".", "", 1).isdigit() else value


def normalize_category(value: str | None, default: str = "plant") -> str:
    candidate = str(value or "").strip().lower()
    return candidate if candidate in VALID_CATEGORIES else default


def category_label(value: str | None) -> str:
    return "자재" if normalize_category(value) == "material" else "식물"


def normalize_purchase_entry_row(
    *,
    category: str,
    name: str,
    vendor: str,
    spec: str,
    quantity: int | float | str,
    purchase_count: int | float | str,
    cost: int | float | str,
    wholesale: int | float | str,
    retail: int | float | str,
) -> tuple[str, str, str, str, float | None, float | None, float | None, float | None, float | None] | None:
    normalized_name = name.strip()
    if not normalized_name:
        return None

    normalized_category = normalize_category(category)
    normalized_quantity = parse_number(quantity)
    normalized_purchase_count = parse_number(purchase_count)
    normalized_cost = parse_number(cost)
    normalized_wholesale = parse_number(wholesale)
    normalized_retail = parse_number(retail)

    return (
        normalized_category,
        normalized_name,
        vendor.strip(),
        normalize_text(spec),
        float(normalized_quantity) if isinstance(normalized_quantity, (int, float)) else None,
        float(normalized_purchase_count) if isinstance(normalized_purchase_count, (int, float)) else None,
        float(normalized_cost) if isinstance(normalized_cost, (int, float)) else None,
        float(normalized_wholesale) if isinstance(normalized_wholesale, (int, float)) else None,
        float(normalized_retail) if isinstance(normalized_retail, (int, float)) else None,
    )


def fetch_purchase_entry(connection: Any, entry_id: int) -> dict[str, Any] | None:
    cursor = connection.execute(
        """
        SELECT id,
               category,
               name,
               vendor,
               spec,
               quantity,
               purchase_count,
               cost,
               wholesale,
               retail,
               date(datetime(created_at, '+9 hours')) AS created_date
        FROM purchase_entries
        WHERE id = ?
        """,
        (entry_id,),
    )
    return cursor_to_dict(cursor)


def fetch_purchase_entries_by_ids(connection: Any, entry_ids: list[int]) -> list[dict[str, Any]]:
    if not entry_ids:
        return []

    placeholders = ", ".join("?" for _ in entry_ids)
    cursor = connection.execute(
        f"""
        SELECT id,
               date(datetime(created_at, '+9 hours')) AS created_date,
               category,
               name,
               vendor,
               spec,
               quantity,
               purchase_count,
               cost,
               wholesale,
               retail
        FROM purchase_entries
        WHERE id IN ({placeholders})
        """,
        tuple(entry_ids),
    )
    rows = cursor_to_dicts(cursor)

    row_map = {int(row["id"]): row for row in rows}
    return [row_map[entry_id] for entry_id in entry_ids if entry_id in row_map]


def build_purchase_workbook(rows: list[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plant Labels"

    header = ["식물 이름", "가격", "매수"]
    header = [
        "ID",
        "저장날짜",
        "식물 이름",
        "매입처",
        "규격",
        "수량",
        "매수",
        "단가",
        "도매가",
        "소매가",
    ]
    sheet.append(header)
    sheet.delete_rows(1, 1)
    sheet.append(["식물 이름", "가격", "매수"])

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue

        quantity = parse_number(row.get("purchase_count", row.get("quantity", "")))
        retail = parse_number(row.get("retail", ""))
        export_quantity: int | float | str = quantity
        if isinstance(quantity, float):
            export_quantity = int(round(quantity))

        sheet.append([name, retail, export_quantity])

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 10

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column == 2:
                    cell.number_format = '#,##0"원"'
                else:
                    cell.number_format = "0"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_purchase_ledger_workbook(rows: list[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchase Ledger"

    header = ["ID", "저장날짜", "식물 이름", "규격", "수량", "단가", "도매가", "소매가"]
    sheet.append(header)
    sheet.delete_rows(1, 1)
    sheet.append(
        ["ID", "저장날짜", "식물 이름", "매입처", "규격", "수량", "매수", "단가", "도매가", "소매가"]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(
            [
                row.get("id", ""),
                row.get("created_date", ""),
                row.get("name", ""),
                row.get("vendor", ""),
                row.get("spec", ""),
                parse_number(row.get("quantity", "")),
                parse_number(row.get("purchase_count", "")),
                parse_number(row.get("cost", "")),
                parse_number(row.get("wholesale", "")),
                parse_number(row.get("retail", "")),
            ]
        )

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 10
    sheet.column_dimensions["F"].width = 10
    sheet.column_dimensions["G"].width = 10
    sheet.column_dimensions["H"].width = 14
    sheet.column_dimensions["I"].width = 14
    sheet.column_dimensions["J"].width = 14

    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column in (6, 7):
                    cell.number_format = "0"
                else:
                    cell.number_format = '#,##0"원"'

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_purchase_ledger_workbook(rows: list[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchase Ledger"

    sheet.append(
        [
            "ID",
            "저장날짜",
            "카테고리",
            "품목명",
            "매입처",
            "규격",
            "수량",
            "매수",
            "단가",
            "도매가",
            "소매가",
        ]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(
            [
                row.get("id", ""),
                row.get("created_date", ""),
                category_label(row.get("category")),
                row.get("name", ""),
                row.get("vendor", ""),
                row.get("spec", ""),
                parse_number(row.get("quantity", "")),
                parse_number(row.get("purchase_count", "")),
                parse_number(row.get("cost", "")),
                parse_number(row.get("wholesale", "")),
                parse_number(row.get("retail", "")),
            ]
        )

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 10
    sheet.column_dimensions["G"].width = 10
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 14
    sheet.column_dimensions["J"].width = 14
    sheet.column_dimensions["K"].width = 14

    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column in (7, 8):
                    cell.number_format = "0"
                else:
                    cell.number_format = '#,##0"원"'

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def parse_ocr_rows(lines: list[dict[str, Any]]) -> list[dict[str, int | str]]:
    items: list[dict[str, int | str]] = []
    seen: set[tuple[str, int, int]] = set()

    for line in lines:
        tokens = [normalize_text(token) for token in line["tokens"]]
        tokens = [token for token in tokens if token]
        if len(tokens) < 4:
            continue

        money_positions = [
            (idx, parse_money_token(token))
            for idx, token in enumerate(tokens)
            if parse_money_token(token) is not None
        ]
        if len(money_positions) < 2:
            continue

        first_money_index, unit_price = money_positions[0]
        if unit_price is None:
            continue

        quantity = None
        quantity_index = None
        for idx in range(first_money_index - 1, -1, -1):
            candidate = parse_quantity_token(tokens[idx])
            if candidate is not None:
                quantity = candidate
                quantity_index = idx
                break

        if quantity is None or quantity_index is None:
            continue

        spec = ""
        name_parts: list[str] = []
        for token in tokens[:quantity_index]:
            if SPEC_TOKEN.match(token):
                spec = token
                break
            if is_name_token(token):
                name_parts.append(token)

        name = " ".join(name_parts).strip()
        if not name:
            continue

        if len(name) <= 1:
            continue

        row_key = (name, unit_price, quantity)
        if row_key in seen:
            continue
        seen.add(row_key)

        items.append(
            {
                "name": name,
                "vendor": "",
                "spec": spec,
                "cost": unit_price,
                "wholesale": unit_price,
                "retail": unit_price * 2,
                "quantity": quantity,
                "purchase_count": quantity,
            }
        )

    return items


def paddle_result_to_lines(result: Any) -> list[dict[str, Any]]:
    raw_lines: list[Any]
    if isinstance(result, list) and result and isinstance(result[0], list):
        raw_lines = result[0]
    elif isinstance(result, list):
        raw_lines = result
    else:
        raw_lines = []

    processed: list[dict[str, Any]] = []
    for entry in raw_lines:
        if not isinstance(entry, list) or len(entry) < 2:
            continue

        box, rec = entry[0], entry[1]
        if not isinstance(rec, (list, tuple)) or len(rec) < 2:
            continue

        text = normalize_text(str(rec[0]))
        score = float(rec[1])
        if not text or score < 0.35:
            continue

        if not isinstance(box, list) or len(box) < 4:
            continue

        xs = [int(point[0]) for point in box if isinstance(point, list) and len(point) >= 2]
        ys = [int(point[1]) for point in box if isinstance(point, list) and len(point) >= 2]
        if not xs or not ys:
            continue

        processed.append(
            {
                "text": text,
                "left": min(xs),
                "right": max(xs),
                "top": min(ys),
                "bottom": max(ys),
                "center_y": (min(ys) + max(ys)) // 2,
            }
        )

    processed.sort(key=lambda item: (item["center_y"], item["left"]))

    lines: list[dict[str, Any]] = []
    for item in processed:
        if not lines or abs(item["center_y"] - lines[-1]["center_y"]) > 28:
            lines.append({"center_y": item["center_y"], "tokens": [item["text"]]})
        else:
            lines[-1]["tokens"].append(item["text"])

    return lines


def build_ocr_variants(image: Image.Image) -> list[Image.Image]:
    rgb = image.convert("RGB")
    gray = ImageOps.autocontrast(image.convert("L"))
    threshold = gray.point(lambda value: 255 if value > 180 else 0)
    sharpened = gray.filter(ImageFilter.SHARPEN)
    return [
        rgb,
        ImageOps.autocontrast(rgb),
        ImageOps.colorize(gray, black="black", white="white").convert("RGB"),
        ImageOps.colorize(sharpened, black="black", white="white").convert("RGB"),
        ImageOps.colorize(threshold, black="black", white="white").convert("RGB"),
    ]


def extract_invoice_items(image: Image.Image) -> tuple[list[dict[str, int | str]], int]:
    base = ImageOps.exif_transpose(image)
    rotations = [
        (0, base),
        (90, base.rotate(90, expand=True)),
        (270, base.rotate(270, expand=True)),
        (180, base.rotate(180, expand=True)),
    ]

    best_items: list[dict[str, int | str]] = []
    best_rotation = 0
    best_score = (-1, -1)
    ocr = get_ocr_engine()

    for rotation, rotated_image in rotations:
        for candidate in build_ocr_variants(rotated_image):
            result = ocr.ocr(np.array(candidate), cls=True)
            lines = paddle_result_to_lines(result)
            items = parse_ocr_rows(lines)
            score = (len(items), sum(len(str(item["name"])) for item in items))
            if score > best_score:
                best_items = items
                best_rotation = rotation
                best_score = score

    return best_items, best_rotation


@app.get("/")
async def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/ledger")
async def ledger_page() -> FileResponse:
    return FileResponse(STATIC_DIR / "ledger.html")


@app.get("/api/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/database/settings")
async def get_database_settings() -> JSONResponse:
    return JSONResponse(get_database_settings_payload())


@app.post("/api/database/turso")
async def save_database_turso_settings(payload: TursoConfigPayload) -> JSONResponse:
    active_environment_config = load_environment_turso_config()
    if active_environment_config:
        source_label_map = {
            "dotenv": ".env 파일",
            "process-env": "시스템 환경변수",
            "mixed-env": "시스템 환경변수 + .env",
        }
        source_label = source_label_map.get(active_environment_config.get("source", ""), ".env")
        return JSONResponse(
            {
                "message": f"현재 Turso 연결은 {source_label}로 관리되고 있습니다. .env 값을 수정해 주세요.",
                "settings": get_database_settings_payload(),
            },
            status_code=400,
        )

    existing_config = load_saved_turso_config()
    url = payload.url.strip() or (existing_config["url"] if existing_config else "")
    auth_token = payload.auth_token.strip() or (
        existing_config["auth_token"] if existing_config else ""
    )

    if not url or not auth_token:
        return JSONResponse(
            {"message": "Turso URL과 토큰을 모두 입력해 주세요."},
            status_code=400,
        )
    if not url.startswith("libsql://"):
        return JSONResponse(
            {"message": "Turso URL은 libsql:// 로 시작해야 합니다."},
            status_code=400,
        )

    TURSO_REPLICA_DIR.mkdir(parents=True, exist_ok=True)
    connection = None
    try:
        libsql = require_libsql()
        connection = libsql.connect(
            str(TURSO_REPLICA_PATH),
            sync_url=url,
            auth_token=auth_token,
        )
        maybe_sync_connection(connection)
        ensure_purchase_entries_schema(connection)
        migrated_count = maybe_migrate_local_rows_to_remote(connection)
    except Exception as error:
        close_connection(connection)
        return JSONResponse(
            {"message": f"Turso 연결에 실패했습니다: {error}"},
            status_code=400,
        )
    finally:
        close_connection(connection)

    save_turso_config(url, auth_token)
    migrated_message = (
        f" 로컬 매입장 {migrated_count}개 항목도 함께 옮겼습니다."
        if migrated_count > 0
        else ""
    )
    return JSONResponse(
        {
            "message": f"Turso 연결을 저장했습니다.{migrated_message}",
            "migrated_count": migrated_count,
            "settings": get_database_settings_payload(),
        }
    )


@app.post("/api/database/backup")
async def backup_database() -> JSONResponse:
    try:
        backup_path, mode = create_database_backup()
    except RuntimeError as error:
        return JSONResponse({"message": str(error)}, status_code=500)

    return JSONResponse(
        {
            "message": "데이터베이스 백업을 만들었습니다.",
            "mode": mode,
            "backup_path": str(backup_path),
        }
    )


@app.post("/api/ocr/preview")
async def ocr_preview(file: UploadFile) -> JSONResponse:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix == ".pdf":
        return JSONResponse(
            {
                "filename": file.filename,
                "message": "현재는 이미지 OCR만 지원합니다. PDF 지원은 다음 단계에서 추가할 수 있습니다.",
                "items": [],
            },
            status_code=400,
        )

    try:
        image = Image.open(BytesIO(await file.read()))
    except UnidentifiedImageError:
        return JSONResponse(
            {
                "filename": file.filename,
                "message": "이미지 파일을 읽지 못했습니다.",
                "items": [],
            },
            status_code=400,
        )

    try:
        items, rotation = extract_invoice_items(image)
    except Exception as error:
        return JSONResponse(
            {
                "filename": file.filename,
                "message": f"PaddleOCR 실행에 실패했습니다: {error}",
                "items": [],
            },
            status_code=500,
        )

    message = (
        f"{len(items)}개 품목을 인식했습니다. "
        f"품목명, 단가, 수량만 가져왔고 회전 보정은 {rotation}도로 적용했습니다."
        if items
        else "품목명을 충분히 읽지 못했습니다. 사진을 반듯하게 촬영하거나 대비를 높여 다시 시도해 보세요."
    )

    return JSONResponse(
        {
            "filename": file.filename,
            "message": message,
            "items": items,
        }
    )


@app.post("/api/export/xlsx")
async def export_xlsx(rows: list[PlantRow]) -> Response:
    buffer = build_purchase_workbook([row.model_dump() for row in rows])

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="plant-labels.xlsx"'
        },
    )


@app.post("/api/purchase-ledger")
async def save_purchase_ledger(rows: list[PlantRow]) -> JSONResponse:
    normalized_rows: list[
        tuple[
            str,
            str,
            str,
            str,
            float | None,
            float | None,
            float | None,
            float | None,
            float | None,
        ]
    ] = []

    for row in rows:
        normalized_row = normalize_purchase_entry_row(
            category=row.category,
            name=row.name,
            vendor=row.vendor,
            spec=row.spec,
            quantity=row.quantity,
            purchase_count=row.purchase_count,
            cost=row.cost,
            wholesale=row.wholesale,
            retail=row.retail,
        )
        if normalized_row is not None:
            normalized_rows.append(normalized_row)

    if not normalized_rows:
        return JSONResponse(
            {"message": "저장할 데이터가 없습니다.", "saved_count": 0},
            status_code=400,
        )

    connection = None
    storage_label = str(DB_PATH)
    try:
        connection, mode = open_purchase_connection(force_sync=True, for_write=True)
        connection.executemany(
            """
            INSERT INTO purchase_entries (
                category,
                name,
                vendor,
                spec,
                quantity,
                purchase_count,
                cost,
                wholesale,
                retail
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            normalized_rows,
        )
        connection.commit()
        maybe_sync_connection(connection)
        if mode == "turso":
            config = get_active_turso_config()
            storage_label = f"Turso ({config['url']})" if config else "Turso"
    except RuntimeError as error:
        return JSONResponse({"message": str(error), "saved_count": 0}, status_code=500)
    finally:
        close_connection(connection)

    return JSONResponse(
        {
            "message": f"{len(normalized_rows)}개 항목을 매입장에 저장했습니다.",
            "saved_count": len(normalized_rows),
            "storage_label": storage_label,
        }
    )


@app.put("/api/purchase-ledger/{entry_id}")
async def update_purchase_ledger(entry_id: int, row: PurchaseLedgerRowUpdate) -> JSONResponse:
    normalized_row = normalize_purchase_entry_row(
        category=row.category,
        name=row.name,
        vendor=row.vendor,
        spec=row.spec,
        quantity=row.quantity,
        purchase_count=row.purchase_count,
        cost=row.cost,
        wholesale=row.wholesale,
        retail=row.retail,
    )

    if normalized_row is None:
        return JSONResponse(
            {"message": "?앸Ъ ?대쫫??鍮꾩슱 ?ㅽ깭濡???ν븷 ???놁뒿?덈떎."},
            status_code=400,
        )

    connection = None
    try:
        connection, _mode = open_purchase_connection(force_sync=True, for_write=True)
        existing_item = fetch_purchase_entry(connection, entry_id)
        if existing_item is None:
            return JSONResponse(
                {"message": "?ㅼ젙???μ뿭??李얠쓣 ???놁뒿?덈떎."},
                status_code=404,
            )

        connection.execute(
            """
            UPDATE purchase_entries
            SET category = ?,
                name = ?,
                vendor = ?,
                spec = ?,
                quantity = ?,
                purchase_count = ?,
                cost = ?,
                wholesale = ?,
                retail = ?
            WHERE id = ?
            """,
            (*normalized_row, entry_id),
        )
        connection.commit()
        maybe_sync_connection(connection)

        item = fetch_purchase_entry(connection, entry_id)
    except RuntimeError as error:
        return JSONResponse({"message": str(error)}, status_code=500)
    finally:
        close_connection(connection)

    return JSONResponse(
        {
            "message": "留ㅼ엯 ???μ뿭???섏젙?덉뒿?덈떎.",
            "item": item,
        }
    )


@app.delete("/api/purchase-ledger/{entry_id}")
async def delete_purchase_ledger(entry_id: int) -> JSONResponse:
    connection = None
    try:
        connection, _mode = open_purchase_connection(force_sync=True, for_write=True)
        existing_item = fetch_purchase_entry(connection, entry_id)
        if existing_item is None:
            return JSONResponse(
                {"message": "?곗궘???μ뿭??李얠쓣 ???놁뒿?덈떎."},
                status_code=404,
            )

        connection.execute(
            """
            DELETE FROM purchase_entries
            WHERE id = ?
            """,
            (entry_id,),
        )
        connection.commit()
        maybe_sync_connection(connection)
    except RuntimeError as error:
        return JSONResponse({"message": str(error)}, status_code=500)
    finally:
        close_connection(connection)

    return JSONResponse({"message": "留ㅼ엯 ???μ뿭???곗궘?덉뒿?덈떎."})


def build_purchase_ledger_query(
    name: str | None, category: str | None, sort: str | None, direction: str | None
) -> tuple[str, list[str]]:
    where_clauses: list[str] = []
    params: list[str] = []

    if name:
        where_clauses.append("name LIKE ?")
        params.append(f"%{name.strip()}%")
    if category:
        normalized_category = normalize_category(category, default="")
        if normalized_category:
            where_clauses.append("category = ?")
            params.append(normalized_category)

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    sort_key = (sort or "date").lower()
    sort_direction = "ASC" if (direction or "desc").lower() == "asc" else "DESC"

    if sort_key == "id":
        order_sql = f"ORDER BY id {sort_direction}"
    elif sort_key == "name":
        order_sql = f"ORDER BY name COLLATE NOCASE {sort_direction}, id DESC"
    else:
        order_sql = f"ORDER BY created_at {sort_direction}, id DESC"

    query = (
        """
        SELECT id, category, name, vendor, spec, quantity, purchase_count, cost, wholesale, retail,
               date(datetime(created_at, '+9 hours')) AS created_date
        FROM purchase_entries
        """
        + where_sql
        + " "
        + order_sql
    )
    return query, params


@app.get("/api/purchase-ledger")
async def get_purchase_ledger(
    name: str | None = None,
    category: str | None = None,
    sort: str | None = None,
    direction: str | None = None,
) -> JSONResponse:
    query, params = build_purchase_ledger_query(name, category, sort, direction)

    connection = None
    try:
        connection, _mode = open_purchase_connection(force_sync=False)
        rows = cursor_to_dicts(connection.execute(query, tuple(params)))
    except RuntimeError as error:
        return JSONResponse({"message": str(error), "items": [], "count": 0}, status_code=500)
    finally:
        close_connection(connection)

    return JSONResponse(
        {
            "items": rows,
            "count": len(rows),
        }
    )


@app.get("/api/purchase-ledger/export")
async def export_purchase_ledger(
    name: str | None = None,
    category: str | None = None,
    sort: str | None = None,
    direction: str | None = None,
) -> Response:
    query, params = build_purchase_ledger_query(name, category, sort, direction)

    connection = None
    try:
        connection, _mode = open_purchase_connection(force_sync=True)
        rows = cursor_to_dicts(connection.execute(query, tuple(params)))
    except RuntimeError as error:
        return Response(
            content=str(error).encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=500,
        )
    finally:
        close_connection(connection)

    buffer = build_purchase_ledger_workbook(rows)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="purchase-ledger.xlsx"'
        },
    )


@app.post("/api/purchase-ledger/export/selected")
async def export_selected_purchase_ledger(selection: PurchaseLedgerExportSelection) -> Response:
    entry_ids = [entry_id for entry_id in selection.ids if isinstance(entry_id, int)]
    if not entry_ids:
        return Response(
            content="선택된 항목이 없습니다.".encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=400,
        )

    connection = None
    try:
        connection, _mode = open_purchase_connection(force_sync=True)
        rows = fetch_purchase_entries_by_ids(connection, entry_ids)
    except RuntimeError as error:
        return Response(
            content=str(error).encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=500,
        )
    finally:
        close_connection(connection)

    if not rows:
        return Response(
            content="내보낼 항목을 찾지 못했습니다.".encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=404,
        )

    buffer = build_purchase_ledger_workbook(rows)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="purchase-ledger-selection.xlsx"'
        },
    )


@app.post("/api/purchase-ledger/export/selected-labels")
async def export_selected_purchase_labels(selection: PurchaseLedgerExportSelection) -> Response:
    entry_ids = [entry_id for entry_id in selection.ids if isinstance(entry_id, int)]
    if not entry_ids:
        return Response(
            content="?좏깮????ぉ???놁뒿?덈떎.".encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=400,
        )

    connection = None
    try:
        connection, _mode = open_purchase_connection(force_sync=True)
        rows = fetch_purchase_entries_by_ids(connection, entry_ids)
    except RuntimeError as error:
        return Response(
            content=str(error).encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=500,
        )
    finally:
        close_connection(connection)

    if not rows:
        return Response(
            content="?대낫????ぉ??李얠? 紐삵뻽?듬땲??".encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            status_code=404,
        )

    buffer = build_purchase_workbook(rows)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="plant-labels-selection.xlsx"'
        },
    )
